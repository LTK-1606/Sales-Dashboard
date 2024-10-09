from openpyxl import load_workbook, Workbook
import re
import os
import sys

def sanitize_sheet_name(name):
    # Replace invalid characters with underscores
    sanitized_name = re.sub(r'[\/:*?"<>|]', '_', name)
    # Truncate to 31 characters if necessary
    return sanitized_name[:31]

def main_consolidate_format_data():
    
    if getattr(sys, 'frozen', False):
        # When running as a bundled executable (e.g., PyInstaller)
        script_dir = os.path.dirname(sys.executable)
        
    else:
        # When running as a script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
    # Path to your Excel file
    input_file_path = os.path.join(script_dir, "sales_dashboard (new).xlsx")
    output_file_path = os.path.join(script_dir, "formatted_sales_dashboard (new).xlsx")

    # Load the original workbook
    input_wb1 = load_workbook(input_file_path)

    # Modify each sheet
    for sheet in input_wb1.worksheets:
        sheet['C11'] = 'Conversion_1'

    # Save the modified workbook
    input_wb1.save(output_file_path)

    # Load formatted workbook
    input_wb2 = load_workbook(os.path.join(script_dir, "formatted_sales_dashboard (new).xlsx"))

    # Create a new workbook for consolidated data
    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # Remove the default sheet

    # Column headers (excluding "Week Start" and "Week End")
    column_headers = [
        "New", "Scrap", "Quotation", "Consignment", "Sales", "Coe Renewal",
        "Loan Paperwork", "Consignment Purchase", "Dealer Purchase", "Floor",
        "Purchases", "Insurances", "Total"
    ]

    # Create a dictionary to hold each new sheet by row header name
    sheets_dict = {}

    # Initialize each sheet in the new workbook with headers
    for row_header in input_wb2.active.iter_rows(min_row=2, max_row=15, min_col=3, max_col=3, values_only=True):
        row_header = row_header[0]  # Get the header from the tuple
        if row_header:
            if row_header in sheets_dict:
                row_header = row_header + "_1"
            sanitized_name = sanitize_sheet_name(row_header)
            sheets_dict[row_header] = output_wb.create_sheet(title=sanitized_name)
            sheets_dict[row_header].append(["Date"] + column_headers)

    # Loop through each sheet (week) in the input workbook
    for sheet_name in input_wb2.sheetnames:
        # Get the current sheet
        input_sheet = input_wb2[sheet_name]

        # Extract the date range from the sheet name
        date_range = sheet_name.split(' to ')[0]

        # Extract row headers dynamically from the current sheet
        for row_idx in range(2, input_sheet.max_row + 1):  # Adjust the range as needed
            row_header = input_sheet.cell(row=row_idx, column=3).value  # Get the row header from column 3
            if row_header and row_header in sheets_dict:
                # Get the data for this row across the defined columns
                data_row = []
                for col_idx in range(4, len(column_headers) + 4):  # Start from 4 to skip "Week Start" and "Week End"
                    cell_value = input_sheet.cell(row=row_idx, column=col_idx).value
                    data_row.append(cell_value)
                
                # Append the date and data row to the corresponding sheet
                sheets_dict[row_header].append([date_range] + data_row)

    # Save the consolidated workbook
    print("Excel file saved at consolidated_&_formatted_data (new).xlsx")
    output_file_path = os.path.join(script_dir, "consolidated_&_formatted_data (new).xlsx")
    output_wb.save(output_file_path)

if __name__ == "__main__":
    main_consolidate_format_data()
