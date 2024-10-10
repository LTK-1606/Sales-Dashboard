import pandas as pd
from openpyxl import load_workbook
import os
import sys

def combine_excel_files(file1, file2, output_file):
    # Load the Excel files
    xl1 = pd.ExcelFile(file1)
    xl2 = pd.ExcelFile(file2)

    # Get the sheet names from both files
    sheets1 = xl1.sheet_names
    sheets2 = xl2.sheet_names

    # Create a writer object for the output file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Iterate through sheets in xl1 (file1)
        for sheet in sheets1:
            df1 = xl1.parse(sheet)

            # Check if the sheet exists in xl2 (file2)
            if sheet in sheets2:
                df2 = xl2.parse(sheet)
            else:
                # If the sheet doesn't exist in xl2, create an empty dataframe
                df2 = pd.DataFrame()

            # Skip empty sheets
            if df1.empty and df2.empty:
                print(f"Skipping empty sheet: {sheet}")
                continue

            # Check if the last row of df1 and the first row of df2 have the same value in the first cell
            if not df1.empty and not df2.empty:
                if df1.iloc[-1, 0] == df2.iloc[0, 0]:
                    # If they match, drop the last row of df1
                    df1 = df1.iloc[:-1]

            # Concatenate the dataframes
            combined_df = pd.concat([df1, df2], ignore_index=True)

            # Write the combined dataframe to the output file
            combined_df.to_excel(writer, sheet_name=sheet, index=False)

    # Update file1 with the combined data
    # Load the workbook for file1
    wb = load_workbook(file1)
    # Remove all existing sheets in file1
    for sheet in wb.sheetnames:
        del wb[sheet]

    # Load the workbook for the output file
    combined_wb = load_workbook(output_file)
    # Copy sheets from the combined workbook to file1
    for sheet_name in combined_wb.sheetnames:
        source_sheet = combined_wb[sheet_name]
        target_sheet = wb.create_sheet(sheet_name)

        for row in source_sheet.iter_rows(values_only=True):
            target_sheet.append(row)

    # Save file1 with updated data
    wb.save(file1)

    print(f"Combined Excel file saved as {output_file} and updated in {file1}")

def main_combine_data():
    if getattr(sys, 'frozen', False):
        # When running as a bundled executable (e.g., PyInstaller)
        script_dir = os.path.dirname(sys.executable)

    else:
        # When running as a script
        script_dir = os.path.dirname(os.path.abspath(__file__))

    file1 = os.path.join(script_dir, "consolidated_&_formatted_data (historical).xlsx")
    file2 = os.path.join(script_dir, "consolidated_&_formatted_data (new).xlsx")
    output_file = os.path.join(script_dir, "cleaned_consolidated_data.xlsx")

    combine_excel_files(file1, file2, output_file)

if __name__ == "__main__":
    main_combine_data()
